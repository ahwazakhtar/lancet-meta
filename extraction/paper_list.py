"""
Load the canonical paper list from `base-data/field and paper list.xlsx`.

Used by the pipeline to:
- enrich extracted papers with bibliographic metadata (title, journal, etc.)
  from the manually curated list
- match PDFs to xlsx rows by DOI

The xlsx is the source of truth for title/authors/year/journal. Claude only
fills in the fields it can derive from the PDF (intervention, design, effect
sizes, etc.).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class PaperListEntry:
    title: str
    authors: str
    year: str
    journal: str
    volume: str
    issue: str
    pages: str
    doi: Optional[str]
    study_id: str  # the "Author Year" label, e.g. "Abdallah 2021"
    tags: str


def _norm_doi(doi: Optional[str]) -> Optional[str]:
    if not doi:
        return None
    s = str(doi).strip()
    # Strip a leading "https://doi.org/" or "doi:" prefix.
    s = re.sub(r"^(?:https?://(?:dx\.)?doi\.org/|doi:\s*)", "", s, flags=re.IGNORECASE)
    return s.lower() or None


def study_id_to_unique_id(study_id: Optional[str]) -> Optional[str]:
    """'Abdallah 2021' -> 'Abdallah2021'."""
    if not study_id:
        return None
    return re.sub(r"\s+", "", str(study_id).strip())


def load_paper_list(xlsx_path: Path) -> tuple[dict[str, PaperListEntry], list[PaperListEntry]]:
    """Returns (by_doi_lowercased, all_entries)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "All studies to extract" not in wb.sheetnames:
        raise RuntimeError(f"Sheet 'All studies to extract' missing from {xlsx_path}")
    ws = wb["All studies to extract"]

    entries: list[PaperListEntry] = []
    by_doi: dict[str, PaperListEntry] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: title, authors, year, month, journal, vol, iss, pages, doi, study, tags, ...
        if not row or not row[0]:
            continue
        title, authors, year, _month, journal, vol, iss, pages, doi, study, tags, *_rest = row
        norm = _norm_doi(doi)
        entry = PaperListEntry(
            title=str(title or "").strip(),
            authors=str(authors or "").strip(),
            year=str(year or "").strip(),
            journal=str(journal or "").strip(),
            volume=str(vol or "").strip(),
            issue=str(iss or "").strip(),
            pages=str(pages or "").strip(),
            doi=norm,
            study_id=str(study or "").strip(),
            tags=str(tags or "").strip(),
        )
        entries.append(entry)
        if norm:
            by_doi[norm] = entry
    logger.info("Loaded %d papers from %s (%d have DOIs)", len(entries), xlsx_path, len(by_doi))
    return by_doi, entries


def lookup_by_doi(by_doi: dict[str, PaperListEntry], doi: str) -> Optional[PaperListEntry]:
    norm = _norm_doi(doi)
    if not norm:
        return None
    return by_doi.get(norm)
